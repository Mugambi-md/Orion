from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font
import platform
import datetime
import os


class ExportReports:
    def export_reports_to_pdf(self, filepath, **sections):
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title_style = styles['Heading2']
        normal_style = styles['Normal']
        export_date = datetime.datetime.now().strftime('%Y/%B/%d %H:%M')
        elements.append(
            Paragraph(f"Exported on: {export_date}", normal_style)
        )
        elements.append(Spacer(1, 12))
        for section_name, data in sections.items():
            if not data:
                continue
            # Section title
            elements.append(Paragraph(section_name.replace("_", " ").title(), title_style))
            elements.append(Spacer(1, 6))
            # Generate headers from keys
            headers = ["No."] + list(data[0].keys())
            table_data = [headers]

            for i, row in enumerate(data):
                row_data = [i + 1] + [row.get(col, "") for col in data[0].keys()]
                table_data.append(row_data)
            bg_color = colors.lightblue if "order" in section_name.lower() else colors.lightgreen
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), bg_color),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            elements.append(table)
            elements.append(PageBreak())
        doc.build(elements)

class ExportReportToExcel:
    def export_to_excel(self, filepath, **sections):
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        bold_font = Font(bold=True)
        ws.append([f"Exported on: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"])
        ws.append([])
        col_widths = {}
        for section_name, records in sections.items():
            # section title
            ws.append([section_name.replace("_", " ").title()])
            ws.cell(row=ws.max_row, column=1).font = bold_font # Make title bold
            ws.append([])
            if records:
                headers = list(records[0].keys())
                ws.append(headers)
                header_row_index = ws.max_row
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row_index, column=col_num)
                    cell.font = bold_font # Make header bold
                    col_widths[col_num] = max(col_widths.get(col_num, 0), len(str(header)) + 2)
                for row in records:
                    values = [str(row.get(h, "")) for h in headers]
                    ws.append(values)
                    for col_num, value in enumerate(values, 1):
                        col_widths[col_num] = max(col_widths.get(col_num, 0), len(value) + 2)
            else:
                ws.append(["No data available."])
            ws.append([])
            ws.append([])
        wb.save(filepath)

class DeliveryExporter:
    def __init__(self, order_id, table_date, status, save_path=None):
        self.order_id = order_id
        self.status = status
        self.table_data = table_date
        self.save_path = save_path or f"Delivery Note Order {self.order_id}.pdf"

    def export_to_pdf(self):
        c = canvas.Canvas(self.save_path, pagesize=A4)
        width, height = A4
        y = height - 50
        c.setFont("Helvetica-Bold", 15)
        c.drawCentredString(width / 2, y, "DELIVERY NOTE.")
        y -= 30
        c.setFont("Helvetica", 12)
        c.drawString(50, y, f"Order No. {self.order_id}. {self.status}")
        y -= 20
        c.drawString(50, y, "-----------------------------------------")
        y -= 20
        headers = ["No.", "Product Code", "Product Name", "Quantity", "Unit Price", "Total Price"]
        col_widths = [40, 80, 150, 60, 70, 80]
        y = height - 130
        c.setFont("Helvetica-Bold", 11)
        for i, header in enumerate(headers):
            c.drawString(50 + sum(col_widths[:i]), y, header)
        y -= 20
        c.setFont("Helvetica", 10)
        for row in self.table_data:
            for i, cell in enumerate(row):
                c.drawString(50 + sum(col_widths[:i]), y, str(cell))
            y -= 20
            if y < 100:
                c.showPage()
                y = height - 50
        c.save()
        return self.save_path
    def print_note(self):
        pdf_path = self.export_to_pdf()
        try:
            system_name = platform.system()
            if system_name == "Windows":
                os.startfile(pdf_path, "print")
            elif system_name == "Darwin": # macOS
                os.system(f"lpr {pdf_path}")
            elif system_name == "Linux":
                os.system(f"lp {pdf_path}")
            else:
                return f"Unsupported Os: {system_name}"
            return "Report sent to printer successfully."
        except Exception as e:
            return f"Printing failed: {e}"
        